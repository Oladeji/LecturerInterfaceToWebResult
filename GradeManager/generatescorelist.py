import openpyxl
import datetime
import base64
import json
from django.conf  import settings

from Crypto.Cipher import AES  #install PyCrypto

from . ScoreSheetClass import ScoreSheetClass
from . basicunit import basicunit 


def returnreporttypes():
      

      return Reports


def validatelist(Scoreslist):
  err =1
  msg = ""
  for score in Scoreslist:
        if score.myScore == -1 :
              err = -1
              msg = "Could Not Process Upload,"+ score.myStudentId +" Has Entries with No Score"
              return(err,msg)
  return(err,msg)

def generatescorelist(excel_file):
            wb = openpyxl.Workbook()        
            wb = openpyxl.load_workbook(excel_file)
            ws = wb["SCORESSHEET"]
            ws.protection.sheet = True
            row_start=7
            try :
                        SECRETKEYciphertext =ws.cell(1, 1).value
                        print(SECRETKEYciphertext)
                        cipher = AES.new(settings.CIPHER_PASS.rjust(32, 'X'),AES.MODE_ECB) # never use ECB in strong systems obviously
                        decodedSecretkey = cipher.decrypt(base64.b64decode(SECRETKEYciphertext)).decode("utf-8") 
                        print(decodedSecretkey)
                        xindex=decodedSecretkey.find('{"CCODE":') 
                        decodedSecretkey= decodedSecretkey[xindex :]
                        Secretkeyobj=json.loads(decodedSecretkey)

          
            except  Exception as inst:
                  return {},{}
                 
              



            myCampId=Secretkeyobj['CID']
            myFacId="XXX"
            myDeptId="XXX"
            myProgId="XXX"
            myProgOptionId="XXX"
            myAsetId = Secretkeyobj['AsetId']
            myAsessionId = Secretkeyobj['AsessionId']
            mySemesterId = Secretkeyobj['SemesterId']
            myLevelTodo = Secretkeyobj['LevelToDo']
            myCourseId = Secretkeyobj['CourseId']
            myCourseUnit = Secretkeyobj['CourseUnit']
            basicunits = basicunit( myCampId,myFacId,myDeptId,myProgId,myProgOptionId,myAsetId,myAsessionId,mySemesterId,myLevelTodo,myCourseId)
            
            total =int(Secretkeyobj['TOTAL'])
            scorelist = []   
            for i in range(row_start,total+ row_start):
                  CourseId =ws.cell(i, 8).value
                  myScoreSheetClassId = ws.cell(i, 9).value
                 
                  StudentId = ws.cell(i, 2).value
                  AsessionId =Secretkeyobj['AsessionId']
                  SemesterId =Secretkeyobj['SemesterId']
                
                  AsetId =Secretkeyobj['AsetId']
                  LevelToDo=Secretkeyobj['LevelToDo']
                  CourseState=ws.cell(i, 10).value
                  CourseUnit  = Secretkeyobj['CourseUnit']
                  CourseNature= ws.cell(i, 11).value
                  Score=ws.cell(i, 6).value 
                  AUserId =1
                  donedate= str(datetime.datetime.today())
                  ReadOnly=ws.cell(i, 13).value
                
                  # d1 = ScoreSheetClass(DetailResultId , StudentId , AsessionId, SemesterId, CourseId,  AsetId,  LevelToDo,  CourseState,  CourseUnit, 
                  # CourseNature,  Score,  AUserId,  donedate,  ReadOnly  ) 
                  # scorelist.append(d1) 

                  d1 = ScoreSheetClass ( myScoreSheetClassId,  "",  "",  "",  StudentId,  AsessionId,  SemesterId,  CourseId,  AsetId,  LevelToDo,  CourseUnit,  CourseNature,  CourseState,  Score,  
                   AUserId,  donedate,  ReadOnly,   ws.cell(i, 12).value,  Secretkeyobj['CID'] ) 
                  scorelist.append(d1) 
                
            return scorelist,basicunits