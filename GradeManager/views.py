from django.shortcuts import render ,redirect
from .forms import UserLoginForm ,UserRegisterForm ,UploadedScoreForm
from django.contrib.auth import  login, authenticate,logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse ,HttpResponseNotFound 
from django.conf  import settings
from django.contrib import messages
import base64
from Crypto.Cipher import AES  #install PyCrypto
from openpyxl.styles import Protection ,Font, Color, Alignment, Border, Side, colors
import openpyxl

import time  
import requests
import json
from . AccessRoles import AccessRoles
from . ScoreSheetClass import ScoreSheetClass
from . generatescorelist import generatescorelist ,validatelist 
from . basicunit import basicunit ,MergebasicScorelist
from django.core.files.storage import FileSystemStorage




def landing(request):
    return render(request,'GradeManager/landing.html')


def logout_view(request):
    return render(request,'Grademanager/logout_view.html')


def login_view(request):
  next = request.GET.get('next')
  if request.method == 'POST':
    form = UserLoginForm(request.POST) 
    if form.is_valid():

       username = form.cleaned_data.get('username')
       password = form.cleaned_data.get('password')
       user = authenticate(username=username,password=password)
       if not user :
               #raise forms.ValidationError('This user does not exist')
               print('This user does not exist 4')
               messages.error(request,"This user does not exist")
               
       if not user.check_password(password):
                  #raise forms.ValidationError('In Correct password')
                  print('This user does not exist 5') 
                  messages.error(request,"In Correct password")
                  
       if not user.is_active:
                  #raise forms.ValidationError('This user is not active')
                   print('This user does not exist 6')
                   messages.error(request,"This user is not active")
                   
       login(request,user)
       if next :
            return redirect(next)

       request.session['serverprogtypeApi'] =request.POST['serverprogtypeApi']
       print(request.session['serverprogtypeApi']) 
       return redirect('landing')
    else:
        for error in form.non_field_errors() :
            messages.error(request,error )  
            form = UserLoginForm()

    context ={'form' : form }
    return render(request,'GradeManager/login_view.html',context)

  context ={'form' : UserLoginForm() }
  return render(request,'GradeManager/login_view.html',context)


def register_view(request):
        print("Calling registerview")
        next = request.GET.get('next')
        print(next)
    #if request.method == 'POST':
        form = UserRegisterForm(request.POST or None) 
        if form.is_valid():
            print("ub4serrr1")
            user = form.save(commit=False)
            print("userrr1")
            Role='USER'
            myCampId=''
            myFacId=''
            myDeptId=''
            myProgId=''
            myProgOptionId=''
            myProgType=''
            myCourseId=''
            theRoles = AccessRoles(Role,myCampId,myFacId,myDeptId,myProgId,myProgOptionId,myProgType,myCourseId)
            #theRolesjson = json.dumps(theRoles)
            print(theRoles)
            user.last_name='theRolesjson'
            user.first_name='Please get this'
            print(user)
            user.save()
            print(user)
            username = form.cleaned_data['email']
            password=form.cleaned_data['password']
            user.username = username
            user.set_password(password)
            print("userrr2")
            user.save()
            print(user)

            new_user = authenticate(username=username,password=password)
            #login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            login(request,new_user)
            if next :
                return redirect(next)
            return redirect('landing')
    #else:
    #    form = UserRegisterForm()
        else:
            print(form.errors)
            for error in form.non_field_errors() :
              messages.error(request,error )  
              print(error)
             
            context ={'form' : form }
            return render(request,'GradeManager/register_view.html',context)


@login_required
def availableCourses_view(request):
    courselist=""
  
    api=settings.BASE_URL+request.session['serverprogtypeApi']+'/api/Camp/PythonGetAvailableCoursesForEmail'
    print(api)
    try:
         
          params={'email':request.user.email}
          
          r = requests.get(api,params)
          courselist = json.loads(r.text)
          messages.success(request, str (len(courselist))+ " Courses Successfully Loaded")
               
    except  Exception as inst:
          print(inst)
          messages.error(request,"Problem Loading Courses , Check Connection")
               
    return render(request,'GradeManager/availableCourses_view.html',{'courselist':courselist,'days':range(1, 32),'months':range(1, 13),})

def processdata(request):  
  
    if request.method=='POST or None':
        print("A Post Message , Details Below")
        print(request.POST)
   
    return render(request,'GradeManager/landing.html')


@login_required
def displayCourse_view(request):  
    courselist={}
    theheader={}
    if request.method=='POST':
        print("A Post Message , Details Below")
        includescoretemp = request.POST.get('includescore','False')
        includescore=False
        if includescoretemp=='on': 
            includescore=True

        orderbymatricnotemp = request.POST.get('orderbymatricno','False')
        orderbymatricno=False
        if orderbymatricnotemp=='on': 
            orderbymatricno=True
        crsid = request.POST['MYCOURSEGUID']
        year = request.POST['year']
        month = request.POST['month']
        day = request.POST['day']
        reportname='Score_Sheet_Printing'
        step='normal'

        params = {'includescore':includescore,'longerreporttype':crsid,'orderbymatricno':orderbymatricno,'reportname':reportname,'step':step,'year':year,'month':month,'day':day}
        api=api=settings.BASE_URL+request.session['serverprogtypeApi']+'/api/Student/PythonPullForscoreEntryUsingCrsGuid'
       
        print(api)
        try:
            headers = {'content-type': 'application/json'}
            r = requests.post(api,json=params,headers=headers)
            courselist = json.loads(r.text)
            theheader={
                    "AsessionId":courselist[0]['MYASESSIONID'],
                    "SemesterId":courselist[0]['MYSEMESTERID'],
                    "LevelToDo":courselist[0]['MYLEVELTODO'],
                    "CourseState":courselist[0]['MYCOURSESTATE'],
                    "CourseUnit":courselist[0]['MYCOURSEUNIT'],
                    "CourseNature":courselist[0]['MYCOURSENATURE'],
                    "AsetId":courselist[0]['MYASETID'],
                    "CourseId":courselist[0]['MYCOURSEID'],

            }
            request.session['courselist'] = courselist
            request.session['params'] = params
            messages.success(request, str (len(courselist))+ " Courses Successfully Loaded")
            
        except  Exception as inst:
            print("See Error Details Below /n")
            print(inst)
            messages.error(request, "Error Check Connection or Contact Admin")
    return render (request,'GradeManager/displayCourse_view.html',{'courselist':courselist})



@login_required
def downloadScoresheet_xls(request):
  courselist={}
  if request.method=='GET':
       if request.session.has_key('courselist'):
                
                courselist = request.session['courselist']
                filename= courselist[0]['MYCOURSEID']+courselist[0]['MYASESSIONID'].replace("/", "_") +courselist[0]['MYSEMESTERID']+'.xls'
                response = HttpResponse(content_type='application/ms-excel')
                response['Content-Disposition'] = "attachment; filename="+filename
           
                workbook = openpyxl.Workbook()
                worksheet=workbook.active
                worksheet.protection.sheet = True
                worksheet.protection.enable()

                bold_font = Font(bold=True)
                big_red_text = Font(bold=True, color="ff0000", size=14)
                center_aligned_text = Alignment(horizontal="center")

               
                workbook.security.set_workbook_password(settings.WORKBOOKHASHED_PASSWORD, already_hashed=False)
                worksheet.protection.password = settings.CIPHER_PASS
                if request.session.has_key('params'):
                    params = request.session['params']
                else:
                    print('params was NOT passed in session')  
                worksheet.title='SCORESSHEET'
                col_start=0
                row_start=7
                
                SECRETKEY={
                    "CCODE":settings.CIPHER_PASS,
                    "TOTAL":str(len(courselist)),
                    "EMAIL":request.user.email,
                    "CID":params['longerreporttype'],
                    "AsessionId":courselist[0]['MYASESSIONID'],
                    "SemesterId":courselist[0]['MYSEMESTERID'],
                    "LevelToDo":courselist[0]['MYLEVELTODO'],
                    "CourseState":courselist[0]['MYCOURSESTATE'],
                    "CourseUnit":courselist[0]['MYCOURSEUNIT'],
                    "CourseNature":courselist[0]['MYCOURSENATURE'],
                    "AsetId":courselist[0]['MYASETID'],
                    "CourseId":courselist[0]['MYCOURSEID'],

                }

                SECRETKEY_STR = json.dumps(SECRETKEY) 
                SECRETKEY_STR= SECRETKEY_STR.rjust(320, 'X')
                cipher = AES.new(settings.CIPHER_PASS.rjust(32, 'X'),AES.MODE_ECB) # never use ECB in strong systems obviously
                SECRETKEYciphertext = base64.b64encode(cipher.encrypt(SECRETKEY_STR))
                print(SECRETKEYciphertext)
                decoded = cipher.decrypt(base64.b64decode(SECRETKEYciphertext))
                worksheet.cell(1 ,1).value=SECRETKEYciphertext
                worksheet.cell(1 ,1).font=Font(color="ffffff", size=2)
               
                worksheet.column_dimensions['B'].width = 20
                worksheet.column_dimensions['C'].width = 20
                worksheet.column_dimensions['D'].width = 20
                worksheet.column_dimensions['E'].width = 20
                now = time.strftime("%x")  
                worksheet.cell(1 ,3).value = now  
                worksheet.cell(1 ,3).font=Font(color="ffffff", size=2)
                worksheet.cell(3 ,2).value='THE POLYTECHNIC IBADAN'
                worksheet.merge_cells('B3:E3')
                worksheet["B3"].font = big_red_text
                worksheet["B3"].alignment = center_aligned_text

                worksheet.cell(4 ,2).value='INTERNAL RESULT DOCUMENT'
                worksheet.merge_cells('B4:E4')
                worksheet["B4"].font = big_red_text
                worksheet["B4"].alignment = center_aligned_text

                worksheet.cell(5 ,2).value='COURSE CODE : '+courselist[0]['MYCOURSEID']+' SESSION : '+courselist[0]['MYASESSIONID'] +' SEMESTER : '+courselist[0]['MYSEMESTERID'] 
                worksheet.merge_cells('B5:E5')

                worksheet["B5"].font = big_red_text
                worksheet["B5"].alignment = center_aligned_text
                
                for index,row in  enumerate(courselist):               
                    worksheet.cell(index+row_start ,col_start+1).value= index+1
                    worksheet.cell(index+row_start ,col_start+1).alignment = center_aligned_text
                    worksheet.cell(index+row_start ,col_start+2).font=bold_font 
                    worksheet.cell(index+row_start ,col_start+2).value= row['MYSTUDENTID']
                    worksheet.cell(index+row_start ,col_start+3).value= row['MYSURNAME']
                    worksheet.cell(index+row_start ,col_start+4).value= row['MYMIDDLENAME']
                    worksheet.cell(index+row_start ,col_start+5 ).value = row['MYFIRSTNAME']

                    worksheet.cell(index+row_start ,col_start+8).value= row['MYCOURSEID']
                    worksheet.cell(index+row_start ,col_start+8).font = Font(color="ffffff", size=2)

                    worksheet.cell(index+row_start ,col_start+9 ).value = row['MYSCORESHEETCLASSID']
                    worksheet.cell(index+row_start ,col_start+9).font = Font(color="ffffff", size=2)

                    worksheet.cell(index+row_start ,col_start+10).value= row['MYCOURSESTATE']
                    worksheet.cell(index+row_start ,col_start+10).font = Font(color="ffffff", size=2)

                    worksheet.cell(index+row_start ,col_start+11 ).value = row['MYCOURSENATURE']
                    worksheet.cell(index+row_start ,col_start+11).font = Font(color="ffffff", size=2)
                    modi= 'False'
                    if row['MYMODIFIED']:
                        modi='True'
                    
                    worksheet.cell(index+row_start ,col_start+12 ).value =modi
                    worksheet.cell(index+row_start ,col_start+12).font = Font(color="ffffff", size=2)

                    worksheet.cell(index+row_start ,col_start+13 ).value =row['MYREADONLY']
                    worksheet.cell(index+row_start ,col_start+13).font = Font(color="ffffff", size=2)




                    

                    worksheet.cell(index+row_start ,col_start+6).value= row['MYSCORE']
                    worksheet.cell(index+row_start ,col_start+6).font=bold_font 
                    worksheet.cell(index+row_start ,col_start+6).alignment = center_aligned_text
                    worksheet.cell(index+row_start ,col_start+6).protection = Protection(locked=False)  
              
                worksheet.cell( len(courselist)+row_start  ,1).value='END'
                worksheet.cell(len(courselist)+row_start  ,1).font = Font(color="ffffff", size=2)
                worksheet.cell(len(courselist)+row_start  ,1).alignment = center_aligned_text
                workbook.save(response)
                return response
       else:
                print('Nothing was passed in session')
  else :
        print('This is a POST message')
        return render (request,'GradeManager/displayCourse_view.html',{'courselist':courselist})

@login_required
def  uploadScoresheet_xls(request):
    context={}
    if request.method == 'POST' :
        
        form = UploadedScoreForm(request.POST,request.FILES)
        if form.is_valid():
            
            excel_file = request.FILES['scoresheetfile']
            thename = (request.user.username).replace('@','').replace('/','')
            fs = FileSystemStorage()
            print(thename)
            filename = fs.save(thename+'-'+excel_file.name, excel_file)
            uploaded_file_url = fs.url(filename)



            try :
                lists,basicunits = generatescorelist(excel_file)


                
                err,msg = validatelist(lists)
                if err == -1 :
                    messages.error(request, msg)
                    return render(request,'GradeManager/uploadScoresheet_xls.html',context)
                #json_string = json.dumps([ob.__dict__ for ob in lists])
                json_string = [ob.__dict__ for ob in lists]
            
                api=settings.BASE_URL+request.session['serverprogtypeApi']+'/api/Student/PythonUploadScore'
                print(api)
                data = MergebasicScorelist(basicunits.__dict__,json_string)
                mydata = json.dumps(data.__dict__)
            
            except  Exception as inst:
               print(excel_file.name+'  Has Error below ') 
               print(inst) 
               messages.error(request, "Problem With the Excel Score File Uploaded")
               return render(request,'GradeManager/uploadScoresheet_xls.html',context)
                 

        try:
            headers = {'content-type': 'application/json'}
            
            r = requests.post(api,data=mydata,headers=headers)
            
            
            if r.status_code==200:
                messages.success(request, "  Successfully Uploaded")
            else:
                messages.error(request,str(r.status_code) +" Problem loading data")
        except  Exception as inst:
            print("See Error Details Below /n")
            print(inst)
            messages.error(request,inst)


    else :
        form = UploadedScoreForm()
    context={'form':form}
    return render(request,'GradeManager/uploadScoresheet_xls.html',context)

@login_required
def downloadScoreSheetPdf(request):

    print('Beginning file download with requests')
    api=settings.BASE_URL+request.session['serverprogtypeApi']+'/api/Student/PythonPullForScoreSheetPdf'
    print(api)
    params = request.session['params']
    headers = {'content-type': 'application/json'}
    r = requests.post(api,json=params,headers=headers)
    response = HttpResponse(r.content,content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename='+params['longerreporttype']+".pdf"
    return response
    #return HttpResponseNotFound('Nothing Found')